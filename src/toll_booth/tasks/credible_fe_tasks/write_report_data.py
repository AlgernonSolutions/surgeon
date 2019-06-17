from aws_xray_sdk.core import xray_recorder


@xray_recorder.capture('write_report_data')
def write_report_data(**kwargs):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    from toll_booth.obj import ObjectDownloadLink
    import os
    from datetime import datetime

    report_bucket_name = kwargs.get('report_bucket_name', os.environ['ASSET_BUCKET'])
    local_user_directory = os.path.expanduser('~')
    report_name = kwargs['report_name']
    id_source = kwargs['id_source']
    today = datetime.utcnow()
    today_string = today.strftime('%Y%m%d')
    report_name = f'{report_name}_{today_string}.xlsx'
    report_save_path = os.path.join(local_user_directory, report_name)
    reports = kwargs['report_data']
    report_book = Workbook()
    front_sheet = report_book.active
    for entry_name, report_data in reports.items():
        new_sheet = report_book.create_sheet(entry_name)
        new_sheet.append([entry_name])
        top_row = new_sheet.row_dimensions[1]
        top_row.font = Font(bold=True, size=18)
        new_sheet.append([])
        row_lengths = [len(x) for x in report_data]
        max_row_length = None
        if row_lengths:
            max_row_length = max([len(x) for x in report_data])
        for row in report_data:
            new_sheet.append(row)
        if max_row_length:
            new_sheet.merge_cells(f'A1:{chr(ord("a") + (max_row_length-1))}1')
    report_book.remove(front_sheet)
    try:
        report_book.save(report_save_path)
    except FileNotFoundError:
        report_save_path = f'/tmp/{report_name}'
        report_book.save(report_save_path)
    download_link = ObjectDownloadLink(
        report_bucket_name, f'{id_source}/{report_name}', local_file_path=report_save_path)
    return {'download_link': download_link}
